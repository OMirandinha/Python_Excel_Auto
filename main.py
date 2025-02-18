from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def create_grades_report(input_file, sheet_name, output_file):
    # Load the workbook and sheet
    wb = load_workbook(input_file)
    ws = wb[sheet_name]

    # Assume the data is in the sheet and starts from row 2 (to avoid headers) and columns A-D
    data = {}

    # Read the data from the sheet (starting from row 2, assuming the first row is the header)
    for row in range(2, ws.max_row + 1):  # Loop through rows starting from 2 (to skip headers)
        name = ws[f"A{row}"].value  # Get student name from column A
        math = ws[f"B{row}"].value  # Get math grade from column B
        science = ws[f"C{row}"].value  # Get science grade from column C
        english = ws[f"D{row}"].value  # Get english grade from column D
        gym = ws[f"E{row}"].value  # Get gym grade from column E

        # Add the student and their grades to the data dictionary
        data[name] = {"math": math, "science": science, "english": english, "gym": gym}

    # Create a new workbook for the report
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = "Grades"

    # Write the headings
    headings = ['Name'] + list(data['Joe'].keys())
    ws_new.append(headings)

    # Write the data
    for person in data:
        grades = list(data[person].values())
        ws_new.append([person] + grades)

    # Calculate the averages for each subject
    for col in range(2, len(data['Joe']) + 2):
        char = get_column_letter(col)
        ws_new[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

    # Make the first row bold and blue
    for col in range(1, 6):
        ws_new[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

    # Save the new workbook with the results
    wb_new.save(output_file)

# Example of how to call the function
input_file = "existing_grades.xlsx"  # Replace this with your input Excel file
sheet_name = "Sheet1"  # Replace this with your sheet name
output_file = "NewGradesReport.xlsx"  # The name of the output file

create_grades_report(input_file, sheet_name, output_file)
